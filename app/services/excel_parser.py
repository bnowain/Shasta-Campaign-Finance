"""Parse NetFile portal Excel bulk exports into row dicts.

The portal exports .xlsx files with one row per transaction/schedule item.
Column names follow the CAL format conventions. This module maps them to
fields compatible with our Transaction, Filing, and Filer models.

Reference: https://public.netfile.com/pub2/docs/Export_Column_Key.xls
"""

from __future__ import annotations

import io
import logging
import zipfile
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# ─── Column mapping ──────────────────────────────────────────
# Maps Excel column names (case-insensitive) to our internal field names.
# Some fields serve double duty depending on the form/schedule type.

_COLUMN_MAP = {
    # Filer identification
    "filer_id": "sos_filer_id",
    "filer_naml": "filer_name",

    # Filing identification
    "report_num": "amendment_seq",       # Amendment sequence (000, 001, 002...)
    "committee_type": "committee_type",  # CTL, BMC, CAO, RCP
    "rpt_date": "filing_date",
    "from_date": "period_start",
    "thru_date": "period_end",

    # Transaction identification
    "tran_id": "tran_id",
    "rec_type": "rec_type",
    "line_item": "line_item",
    "entity_cd": "entity_type",
    "memo_code": "memo_code",
    "form_type": "schedule",             # Actually the schedule letter (A, B, C, E...)

    # Contributor / Payee name fields
    "tran_naml": "tran_last_name",
    "tran_namf": "tran_first_name",
    "tran_namt": "tran_name_title",
    "tran_nams": "tran_name_suffix",

    # Payee fields (for expenditures)
    "payee_naml": "payee_last_name",
    "payee_namf": "payee_first_name",
    "payee_namt": "payee_name_title",
    "payee_nams": "payee_name_suffix",

    # Lender fields (for loans)
    "lndr_naml": "lender_last_name",
    "lndr_namf": "lender_first_name",

    # Address
    "tran_city": "city",
    "tran_st": "state",
    "tran_zip4": "zip_code",

    # Employment
    "tran_emp": "employer",
    "tran_occ": "occupation",

    # Amounts and dates
    "tran_date": "tran_date",
    "tran_amt1": "tran_amount",
    "tran_amt2": "cumulative_amount",
    "amount": "amount",
    "expn_date": "expn_date",
    "expn_code": "expn_code",
    "expn_dscr": "description",
    "tran_dscr": "tran_description",
    "tran_type": "tran_type",            # MON, NON, etc.
}


def _normalize_header(name: str) -> str:
    """Normalize a header name for matching."""
    return name.strip().lower().replace(" ", "_")


def _parse_date(val) -> date | None:
    """Coerce a cell value to a date."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    # Try string parsing
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_float(val) -> float | None:
    """Coerce a cell value to a float."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("$", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_bool(val) -> bool:
    """Coerce a cell value to a boolean (for memo_code)."""
    if val is None:
        return False
    s = str(val).strip().upper()
    return s in ("TRUE", "1", "Y", "YES", "X")


def _resolve_name_fields(row: dict) -> tuple[str | None, str | None, str | None]:
    """Determine the entity name fields from the row.

    Priority: tran_* (contributions) > payee_* (expenditures) > lender_* (loans).
    Returns (last_name, first_name, entity_name).
    """
    last = row.get("tran_last_name") or row.get("payee_last_name") or row.get("lender_last_name")
    first = row.get("tran_first_name") or row.get("payee_first_name") or row.get("lender_first_name")

    entity_name = None
    if last:
        parts = [p for p in [last, first] if p]
        entity_name = ", ".join(parts)

    return last, first, entity_name


def _resolve_amount(row: dict) -> float:
    """Get the best amount value from the row."""
    # tran_amount is the primary field for contributions
    # amount is used for expenditures
    amt = row.get("tran_amount")
    if amt is not None:
        return amt
    amt = row.get("amount")
    if amt is not None:
        return amt
    return 0.0


def _resolve_date(row: dict) -> date | None:
    """Get the best transaction date from the row."""
    return row.get("tran_date") or row.get("expn_date")


def parse_excel_export(path: Path | str) -> list[dict]:
    """Parse a NetFile portal Excel export into a list of row dicts.

    Each dict contains both filer/filing identification fields and
    transaction fields, ready for grouping and upserting.

    Returns list of dicts with keys:
        - sos_filer_id, filer_name (filer identification)
        - report_num, form_type, filing_date, period_start, period_end (filing)
        - last_name, first_name, entity_name, entity_type
        - city, state, zip_code, employer, occupation
        - amount, cumulative_amount, transaction_date
        - tran_id, schedule, transaction_type_code, description, memo_code
        - data_source: "excel_export"
    """
    path = Path(path)
    logger.info("Parsing Excel export: %s", path.name)

    # The portal wraps the xlsx inside a ZIP archive — detect and extract
    file_bytes = path.read_bytes()
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except KeyError:
        # Not a direct xlsx — try extracting from ZIP wrapper
        logger.info("File is ZIP-wrapped, extracting inner xlsx...")
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            xlsx_names = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
            if not xlsx_names:
                raise ValueError(f"No .xlsx file found inside ZIP: {path}")
            inner_name = xlsx_names[0]
            logger.info("Found inner file: %s", inner_name)
            inner_bytes = zf.read(inner_name)
        wb = load_workbook(io.BytesIO(inner_bytes), read_only=True, data_only=True)
    ws = wb.active

    # Auto-detect header row (first row with recognized column names)
    rows = ws.iter_rows()
    header_row = next(rows)
    raw_headers = [str(cell.value or "").strip() for cell in header_row]
    normalized = [_normalize_header(h) for h in raw_headers]

    # Build column index → field name mapping
    col_map: dict[int, str] = {}
    unmapped = []
    for i, norm_name in enumerate(normalized):
        if norm_name in _COLUMN_MAP:
            col_map[i] = _COLUMN_MAP[norm_name]
        elif norm_name:
            unmapped.append(raw_headers[i])

    if unmapped:
        logger.debug("Unmapped columns: %s", ", ".join(unmapped))

    mapped_count = len(col_map)
    logger.info("Mapped %d/%d columns", mapped_count, len([h for h in normalized if h]))

    # Parse data rows
    results = []
    skipped = 0

    for row_cells in rows:
        raw = {}
        for i, cell in enumerate(row_cells):
            if i in col_map:
                raw[col_map[i]] = cell.value

        # Skip completely empty rows
        if not any(v is not None for v in raw.values()):
            continue

        # Parse typed fields
        sos_filer_id = str(raw.get("sos_filer_id", "") or "").strip()
        filing_date = _parse_date(raw.get("filing_date"))
        amendment_seq = str(raw.get("amendment_seq", "") or "").strip()

        # Build composite filing key: FilerID_RptDate_AmendSeq
        filing_date_str = filing_date.isoformat() if filing_date else "unknown"
        filing_key = f"{sos_filer_id}_{filing_date_str}_{amendment_seq}"

        # Resolve description from tran_description or expn description
        description = str(raw.get("description", "") or "").strip()
        if not description:
            description = str(raw.get("tran_description", "") or "").strip()

        parsed = {
            "sos_filer_id": sos_filer_id,
            "filer_name": str(raw.get("filer_name", "") or "").strip(),
            "filing_key": filing_key,
            "amendment_seq": amendment_seq,
            "committee_type": str(raw.get("committee_type", "") or "").strip(),
            "filing_date": filing_date,
            "period_start": _parse_date(raw.get("period_start")),
            "period_end": _parse_date(raw.get("period_end")),
            "tran_id": str(raw.get("tran_id", "") or "").strip(),
            "rec_type": str(raw.get("rec_type", "") or "").strip(),
            "entity_type": str(raw.get("entity_type", "") or "").strip(),
            "city": str(raw.get("city", "") or "").strip(),
            "state": str(raw.get("state", "") or "").strip(),
            "zip_code": str(raw.get("zip_code", "") or "").strip(),
            "employer": str(raw.get("employer", "") or "").strip(),
            "occupation": str(raw.get("occupation", "") or "").strip(),
            "description": description,
            "schedule": str(raw.get("schedule", "") or "").strip(),
            "transaction_type_code": str(raw.get("expn_code", "") or "").strip(),
            "tran_type": str(raw.get("tran_type", "") or "").strip(),
            "memo_code": _parse_bool(raw.get("memo_code")),
            "data_source": "excel_export",
        }

        # Parse name fields from raw values
        for key in ("tran_last_name", "tran_first_name", "payee_last_name",
                     "payee_first_name", "lender_last_name", "lender_first_name"):
            parsed[key] = str(raw.get(key, "") or "").strip()

        # Parse amounts
        parsed["tran_amount"] = _parse_float(raw.get("tran_amount"))
        parsed["cumulative_amount"] = _parse_float(raw.get("cumulative_amount"))
        parsed["amount_raw"] = _parse_float(raw.get("amount"))

        # Parse dates
        parsed["tran_date"] = _parse_date(raw.get("tran_date"))
        parsed["expn_date"] = _parse_date(raw.get("expn_date"))

        # Resolve composite fields
        last_name, first_name, entity_name = _resolve_name_fields(parsed)
        parsed["last_name"] = last_name or ""
        parsed["first_name"] = first_name or ""
        parsed["entity_name"] = entity_name or ""
        parsed["amount"] = _resolve_amount(parsed)
        parsed["transaction_date"] = _resolve_date(parsed)

        # Skip rows with zero/null amount (header repeats, summary rows, etc.)
        if parsed["amount"] == 0.0 and not parsed["tran_id"]:
            skipped += 1
            continue

        results.append(parsed)

    wb.close()
    logger.info("Parsed %d transactions (%d skipped) from %s", len(results), skipped, path.name)
    return results
