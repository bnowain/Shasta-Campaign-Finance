"""Parse Clarity Elections result files (CSV + Excel) into unified race results.

Supports three formats:
  A) "Detailed vote totals" CSV (2022-2023) — one row per choice per precinct
  B1) Single-sheet cumulative Excel (2024 Primary, 2020, 2021) — stacked contest blocks
  B2) Multi-sheet canvass Excel (2024 General) — one sheet per contest with precinct rows
"""

from __future__ import annotations

import csv
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

logger = logging.getLogger("clarity_parser")


# ─── Unified output dataclasses ──────────────────────────────

@dataclass
class ClarityCandidateResult:
    name: str                   # "DARIN HALE"
    party: str | None = None
    votes: int = 0
    vote_pct: float | None = None
    is_writein: bool = False    # "(W)" suffix or "write-in" variants


@dataclass
class ClarityRaceResult:
    contest_name: str           # "Supervisor, District 3"
    is_measure: bool = False    # True for "Measure A", "Prop 1", etc.
    candidates: list[ClarityCandidateResult] = field(default_factory=list)
    registered_voters: int | None = None
    ballots_cast: int | None = None
    turnout_pct: float | None = None


# ─── Auto-detect + dispatch ──────────────────────────────────

def parse_file(path: Path) -> list[ClarityRaceResult]:
    """Auto-detect format and parse a Clarity result file."""
    suffix = path.suffix.lower()

    if suffix == '.csv':
        # Check if this is a CVR (Cast Vote Record) file
        if _is_cvr_csv(path):
            return _parse_cvr_csv(path)
        return parse_csv(path)
    elif suffix == '.xlsx':
        return parse_excel(path)
    else:
        logger.warning("Unsupported file type: %s", suffix)
        return []


def _is_cvr_csv(path: Path) -> bool:
    """Check if a CSV file is a CVR (Cast Vote Record) ballot-level export."""
    try:
        with open(path, 'r', encoding='utf-8-sig') as f:
            first_line = f.readline()
            # CVR files have contest names with "(Vote For=N)" in the first row
            if '(Vote For=' in first_line or 'CvrNumber' in first_line:
                return True
            # Check second line for candidate names pattern
            second_line = f.readline()
            if second_line and first_line.count(',') > 20:
                # Many columns + no standard header = likely CVR
                headers = first_line.strip().split(',')
                if not any(h.strip() in ('Contest Title', 'CONTEST_FULL_NAME', '#Precinct', 'Precinct') for h in headers[:10]):
                    return True
    except Exception:
        pass
    return False


def parse_excel(path: Path) -> list[ClarityRaceResult]:
    """Auto-detect Excel sub-format and parse."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    try:
        if "Count_CumulativeReport" in sheet_names:
            # Format E: Cumulative single-sheet (2024 Primary style)
            logger.info("Detected format E (cumulative single-sheet): %s", path.name)
            return _parse_cumulative(wb)

        if "Document map" in sheet_names and len(sheet_names) > 5:
            # Check if Format D (2024 canvass) or Format F (district canvass)
            # Format D has contest title at row 25; Format F has it at row 5
            test_ws = wb[sheet_names[1]] if len(sheet_names) > 1 else None
            if test_ws:
                test_rows = list(test_ws.iter_rows(values_only=True, max_row=10))
                if len(test_rows) >= 8:
                    # Format F: row 5 col C has "N ContestName" pattern
                    row5_c = str(test_rows[5][2]).strip() if len(test_rows[5]) > 2 and test_rows[5][2] else ""
                    if re.match(r'^\d+\s+', row5_c):
                        logger.info("Detected format F (district-total canvass): %s", path.name)
                        return _parse_district_canvass(wb)

            logger.info("Detected format D (multi-sheet canvass): %s", path.name)
            return _parse_canvass(wb)

        # Check for Format G/H: named contest sheets or SheetN with contest titles
        if _detect_sov_format(wb, sheet_names):
            logger.info("Detected format G/H (SOV by-district/precinct): %s", path.name)
            return _parse_sov(wb)

        logger.info("Attempting generic Excel parse: %s (sheets: %s)", path.name, sheet_names)
        return _parse_generic_excel(wb)
    finally:
        wb.close()


def _detect_sov_format(wb: openpyxl.Workbook, sheet_names: list[str]) -> bool:
    """Detect SOV (Statement of Votes) format G or H."""
    # Check for named contest sheets (Format G: DemPP, RepPP, Prop13, etc.)
    known_contest_names = {'DemPP', 'RepPP', 'AIPP', 'GRNPP', 'LIBPP', 'PFPP',
                           'Gov', 'Lt Gov', 'SoS', 'Controller', 'Treasurer', 'AG'}
    if any(sn in known_contest_names for sn in sheet_names):
        return True

    # Check for SheetN with "Vote for" contest titles (Format H)
    for sn in sheet_names[1:3]:  # Check Sheet2/Sheet3
        try:
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True, max_row=4))
            if len(rows) >= 2:
                row1 = str(rows[1][0]).strip() if rows[1][0] else ""
                if 'Vote for' in row1 or 'Vote For' in row1:
                    return True
        except Exception:
            pass
    return False


# ─── Parser A: CSV "Detailed vote totals" ────────────────────

def parse_csv(path: Path) -> list[ClarityRaceResult]:
    """Parse Clarity CSV formats (Detailed vote totals + Generic ENR Export).

    Supports two column naming conventions:
      - Format A: Contest Title, Choice Name, Total Votes, Registered Voters, Ballots Cast, Choice Party
      - Format B: CONTEST_FULL_NAME, CANDIDATE_FULL_NAME, VOTE_COUNT, REGISTER_VOTERS, BALLOT_CAST, PARTY_CODE
    """
    logger.info("Parsing CSV: %s", path.name)

    # Column alias mapping: Format B -> Format A canonical names
    COLUMN_ALIASES = {
        "CONTEST_FULL_NAME": "Contest Title",
        "CANDIDATE_FULL_NAME": "Choice Name",
        "VOTE_COUNT": "Total Votes",
        "REGISTER_VOTERS": "Registered Voters",
        "BALLOT_CAST": "Ballots Cast",
        "PARTY_CODE": "Choice Party",
    }

    rows: list[dict] = []
    with open(path, 'r', encoding='utf-8-sig') as f:
        # Skip format version line if present
        first_line = f.readline()
        if not first_line.startswith('#FormatVersion'):
            f.seek(0)

        reader = csv.DictReader(f)
        for row in reader:
            # Normalize column names via alias mapping
            normalized = {}
            for key, val in row.items():
                canonical = COLUMN_ALIASES.get(key, key)
                normalized[canonical] = val
            rows.append(normalized)

    if not rows:
        logger.warning("CSV is empty: %s", path.name)
        return []

    # Group by contest
    contests: dict[str, dict] = {}
    for row in rows:
        contest = row.get('Contest Title', '').strip()
        choice = row.get('Choice Name', '').strip()
        if not contest or not choice:
            continue

        if contest not in contests:
            contests[contest] = {
                'choices': {},
                'registered': 0,
                'ballots': 0,
            }

        # Accumulate votes across precincts
        total_votes = _safe_int(row.get('Total Votes', '0'))
        if choice not in contests[contest]['choices']:
            contests[contest]['choices'][choice] = {
                'votes': 0,
                'party': row.get('Choice Party', '').strip() or None,
            }
        contests[contest]['choices'][choice]['votes'] += total_votes

        # Track registered/ballots from first precinct row
        reg = _safe_int(row.get('Registered Voters', '0'))
        bal = _safe_int(row.get('Ballots Cast', '0'))
        contests[contest]['registered'] += reg
        contests[contest]['ballots'] += bal

    # Convert to unified format
    results: list[ClarityRaceResult] = []
    # We need to deduplicate registered/ballots — they're per-precinct-per-choice,
    # so divide by number of choices to get per-precinct totals
    for contest_name, data in contests.items():
        n_choices = max(len(data['choices']), 1)
        registered = data['registered'] // n_choices if data['registered'] else None
        ballots = data['ballots'] // n_choices if data['ballots'] else None
        turnout = (ballots / registered * 100) if registered and ballots else None

        race = ClarityRaceResult(
            contest_name=contest_name,
            is_measure=_is_measure(contest_name),
            registered_voters=registered,
            ballots_cast=ballots,
            turnout_pct=round(turnout, 2) if turnout else None,
        )

        total_votes = sum(c['votes'] for c in data['choices'].values())

        for choice_name, choice_data in data['choices'].items():
            # Skip write-in noise
            if _is_rejected_writein(choice_name):
                continue

            votes = choice_data['votes']
            pct = round(votes / total_votes * 100, 2) if total_votes > 0 else None

            race.candidates.append(ClarityCandidateResult(
                name=_clean_name(choice_name),
                party=choice_data['party'],
                votes=votes,
                vote_pct=pct,
                is_writein=_is_writein(choice_name),
            ))

        # Sort by votes descending
        race.candidates.sort(key=lambda c: -c.votes)
        results.append(race)

    logger.info("CSV parsed: %d contests, %d total candidates",
                len(results), sum(len(r.candidates) for r in results))
    return results


# ─── Parser B1: Single-sheet cumulative ──────────────────────

def _parse_cumulative(wb: openpyxl.Workbook) -> list[ClarityRaceResult]:
    """Parse single-sheet cumulative format (2024 Primary style).

    Structure: stacked contest blocks on 'Count_CumulativeReport' sheet.
    Each block:
      - Contest title row (e.g., "President - Democratic Party - Vote for One")
      - "Precincts Counted" row
      - "Voters" row (ballots/registered)
      - Column header row ("Choice | Party | Vote By Mail | ... | Election Day Voting")
      - Candidate rows (name | party | VBM | % | EDV | % | Total | %)
      - "Cast Votes:" / "Undervotes:" / "Overvotes:" summary rows
    """
    ws = wb["Count_CumulativeReport"]
    rows = list(ws.iter_rows(values_only=True))

    results: list[ClarityRaceResult] = []
    i = 0

    while i < len(rows):
        row = rows[i]
        if not row or len(row) == 0:
            i += 1
            continue
        cell_a = str(row[0]).strip() if row[0] else ""

        # Detect contest title — has " - Vote for " or is a known pattern
        if _is_contest_title(cell_a):
            race, i = _parse_cumulative_block(rows, i)
            if race:
                results.append(race)
            continue

        i += 1

    logger.info("Cumulative parsed: %d contests", len(results))
    return results


def _is_contest_title(text: str) -> bool:
    """Detect if a row is a contest title in cumulative format."""
    if not text:
        return False
    # Contest titles contain " - Vote for " or end with known patterns
    if " - Vote for " in text:
        return True
    if re.match(r'^(Measure|Prop|Proposition)\s+', text, re.IGNORECASE):
        return True
    if re.match(r'^(Recall|District\s+\d)', text, re.IGNORECASE):
        return True
    return False


def _parse_cumulative_block(rows: list, start: int) -> tuple[ClarityRaceResult | None, int]:
    """Parse a single contest block from cumulative sheet starting at 'start'."""
    title_row = rows[start]
    contest_name = str(title_row[0]).strip()

    # Clean up the title — remove " - Vote for One/Two/etc"
    clean_title = re.sub(r'\s*-\s*Vote for\s+\w+\s*$', '', contest_name, flags=re.IGNORECASE)
    # Remove party from title for partisan primaries
    # e.g., "President - Democratic Party" -> keep as-is for clarity
    clean_title = clean_title.strip()

    race = ClarityRaceResult(
        contest_name=clean_title,
        is_measure=_is_measure(clean_title),
    )

    i = start + 1
    in_candidates = False

    while i < len(rows):
        row = rows[i]
        if not row or len(row) == 0:
            i += 1
            continue
        cell_a = str(row[0]).strip() if row[0] else ""

        # Parse voter registration/ballots line
        if cell_a.startswith("Voters"):
            # "Voters" row: col B or nearby has "ballots / registered" or just numbers
            # Try to find numbers in the row
            for cell in row[1:]:
                if cell and isinstance(cell, (int, float)):
                    if not race.ballots_cast:
                        race.ballots_cast = int(cell)
                    elif not race.registered_voters:
                        race.registered_voters = int(cell)
            i += 1
            continue

        # Column header row — marks start of candidate data
        if cell_a == "Choice":
            in_candidates = True
            i += 1
            continue

        # End-of-block markers
        if cell_a in ("Cast Votes:", "Undervotes:", "Overvotes:"):
            i += 1
            # Skip remaining summary rows
            while i < len(rows):
                next_a = str(rows[i][0]).strip() if rows[i][0] else ""
                if next_a in ("Cast Votes:", "Undervotes:", "Overvotes:"):
                    i += 1
                else:
                    break
            break

        # End of report
        if "End of report" in cell_a:
            i += 1
            break

        # New contest title means we're done with this block
        if i > start + 1 and _is_contest_title(cell_a):
            break

        # Parse candidate row
        if in_candidates and cell_a and cell_a not in ("", "Choice"):
            name = cell_a
            party = str(row[1]).strip() if row[1] else None
            if party == "None":
                party = None

            # Find vote totals — scan for numeric columns
            # Typical layout: Name | Party | VBM | %VBM | EDV | %EDV | Total | %Total
            # But column positions vary. Look for the largest numeric values.
            numbers = []
            for j, cell in enumerate(row[2:], start=2):
                if isinstance(cell, (int, float)) and cell == int(cell):
                    numbers.append((j, int(cell)))

            total_votes = 0
            vote_pct = None

            if numbers:
                # The total votes is typically the second-to-last or a large number
                # In cumulative format: VBM, VBM%, EDV, EDV%, Total, Total%
                # We want the "Total" column
                vote_values = [n[1] for n in numbers]

                # Find percentages (values between 0 and 100 that look like %)
                # The total is usually the largest non-percentage value
                for j, cell in enumerate(row[2:], start=2):
                    if isinstance(cell, (int, float)):
                        pass

                # Simpler approach: last integer before last float-like value
                # Actually, let's just find all ints and take the sum of VBM+EDV
                # or the explicit total column
                if len(vote_values) >= 3:
                    # Assume: VBM, EDV, Total (or VBM, EDV with pcts interspersed)
                    total_votes = max(vote_values)  # Total is always >= any single component
                elif len(vote_values) == 1:
                    total_votes = vote_values[0]

            # Look for percentage
            for cell in row[2:]:
                if isinstance(cell, float) and 0 < cell < 1:
                    vote_pct = round(cell * 100, 2)
                elif isinstance(cell, (int, float)) and cell != total_votes:
                    pass

            race.candidates.append(ClarityCandidateResult(
                name=_clean_name(name),
                party=party,
                votes=total_votes,
                vote_pct=vote_pct,
                is_writein=_is_writein(name),
            ))

        i += 1

    # Compute turnout if we have the data
    if race.registered_voters and race.ballots_cast:
        race.turnout_pct = round(race.ballots_cast / race.registered_voters * 100, 2)

    # Sort candidates by votes
    race.candidates.sort(key=lambda c: -c.votes)

    # Compute percentages if missing
    total_cast = sum(c.votes for c in race.candidates)
    for c in race.candidates:
        if c.vote_pct is None and total_cast > 0:
            c.vote_pct = round(c.votes / total_cast * 100, 2)

    return race, i


# ─── Parser B2: Multi-sheet canvass ──────────────────────────

def _parse_canvass(wb: openpyxl.Workbook) -> list[ClarityRaceResult]:
    """Parse multi-sheet canvass format (2024 General style).

    Structure: 'Document map' + one sheet per contest.
    Each contest sheet:
      - Row 25: contest title
      - Row 27: candidate names as column headers (cols D+)
      - Rows 29-96: precinct data
      - Last data row: "Totals"
      - Trailing columns: Cast Votes, Undervotes, Overvotes, VBM Ballots, ED Ballots, Total Ballots, Registered Voters
    """
    results: list[ClarityRaceResult] = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "Document map":
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 28:
            continue

        # Row 25 (index 24): contest title
        contest_row = rows[24] if len(rows) > 24 else None
        contest_name = ""
        if contest_row:
            contest_name = str(contest_row[0]).strip() if contest_row[0] else ""

        if not contest_name:
            # Try scanning nearby rows for the title
            for ri in range(20, min(28, len(rows))):
                cell = rows[ri][0] if rows[ri][0] else ""
                text = str(cell).strip()
                if text and text not in ("Canvass Results", "Canvass Results - Final",
                                         "Presidential General Election", ""):
                    # Skip date-like and metadata rows
                    if not re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', text) and \
                       not re.match(r'^\d{1,2}:\d{2}', text) and \
                       text != "Page" and not text.startswith("Shasta"):
                        contest_name = text
                        break

        if not contest_name:
            logger.debug("Skipping sheet '%s' — no contest title found", sheet_name)
            continue

        # Row 27 (index 26): column headers — candidate names
        header_row = rows[26] if len(rows) > 26 else None
        if not header_row:
            continue

        # Build candidate column mapping
        # Col A = "Precinct", cols D+ = candidate names
        # Some columns are spacers (None)
        candidate_cols: list[tuple[int, str]] = []
        meta_cols: dict[str, int] = {}  # Cast Votes, Undervotes, etc.

        meta_labels = {
            'cast votes', 'undervotes', 'overvotes',
            'vote by mail ballots cast', 'election day voting ballots cast',
            'total ballots cast', 'registered voters',
            'turnout percentage', 'turnout',
        }

        for col_idx, cell in enumerate(header_row):
            if cell is None:
                continue
            text = str(cell).strip()
            if not text or text.lower() == 'precinct':
                continue
            if text.lower() in meta_labels:
                meta_cols[text.lower()] = col_idx
            else:
                candidate_cols.append((col_idx, text))

        if not candidate_cols:
            continue

        # Find totals row — scan from bottom
        totals_row = None
        for ri in range(len(rows) - 1, 26, -1):
            cell_a = str(rows[ri][0]).strip() if rows[ri][0] else ""
            if cell_a.lower() == "totals":
                totals_row = rows[ri]
                break

        if totals_row is None:
            logger.debug("Skipping sheet '%s' — no Totals row", sheet_name)
            continue

        race = ClarityRaceResult(
            contest_name=contest_name,
            is_measure=_is_measure(contest_name),
        )

        # Extract metadata from totals row
        if 'registered voters' in meta_cols:
            val = totals_row[meta_cols['registered voters']]
            race.registered_voters = int(val) if val else None
        if 'total ballots cast' in meta_cols:
            val = totals_row[meta_cols['total ballots cast']]
            race.ballots_cast = int(val) if val else None
        if race.registered_voters and race.ballots_cast:
            race.turnout_pct = round(race.ballots_cast / race.registered_voters * 100, 2)

        # Extract candidate votes from totals row
        total_cast_votes = 0
        for col_idx, cand_name in candidate_cols:
            votes = totals_row[col_idx]
            votes = int(votes) if votes else 0
            total_cast_votes += votes

            # Parse party from name (e.g., "DONALD J. TRUMP, REP")
            name, party = _split_party(cand_name)

            race.candidates.append(ClarityCandidateResult(
                name=_clean_name(name),
                party=party,
                votes=votes,
                is_writein=_is_writein(cand_name),
            ))

        # Compute percentages
        for c in race.candidates:
            if total_cast_votes > 0:
                c.vote_pct = round(c.votes / total_cast_votes * 100, 2)

        # Sort by votes
        race.candidates.sort(key=lambda c: -c.votes)
        results.append(race)

    logger.info("Canvass parsed: %d contests", len(results))
    return results


# ─── Parser F: District-Total Canvass (2020 General, 2021 Recall) ──

def _parse_district_canvass(wb: openpyxl.Workbook) -> list[ClarityRaceResult]:
    """Parse district-total canvass format (2020 General, 2021 Recall style).

    Structure: Document map + per-contest sheets. Key layout:
      - Row 5, col C: contest name(s) with numeric prefix "1 ContestName"
      - Row 6: contest numbering per column (1,1,1... 2,2,2...)
      - Row 7: col C="Registered Voters", col D="Voters Cast", cols G+= candidate names
      - Row 8: "Electionwide" totals row
    Multiple contests can share a sheet (split by numbering in row 6).
    Multi-sheet per contest: merge candidates from consecutive sheets sharing the same contest.
    """
    results_map: dict[str, ClarityRaceResult] = {}  # contest_name -> race

    for sheet_name in wb.sheetnames:
        if sheet_name == "Document map":
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True, max_row=15))

        if len(rows) < 9:
            continue

        # Row 5 (idx 5), col C: contest name(s) with numeric prefix
        row5 = rows[5] if len(rows) > 5 else ()
        contest_raw = str(row5[2]).strip() if len(row5) > 2 and row5[2] else ""
        if not contest_raw or not re.match(r'^\d+\s+', contest_raw):
            continue

        # Row 6: contest numbering per column
        row6 = rows[6] if len(rows) > 6 else ()
        # Row 7: candidate names
        row7 = rows[7] if len(rows) > 7 else ()
        # Row 8: Electionwide totals
        row8 = rows[8] if len(rows) > 8 else ()

        if not row7 or not row8:
            continue

        # Parse contest numbering from row 6 to group columns by contest
        # Columns G (6) onward may have contest numbers: 1,1,1... 2,2,2...
        contest_nums: dict[int, list[int]] = {}  # contest_num -> [col_indices]
        for col_idx in range(6, len(row6)):
            cell = row6[col_idx] if col_idx < len(row6) else None
            if cell is not None:
                try:
                    num = int(cell)
                    contest_nums.setdefault(num, []).append(col_idx)
                except (ValueError, TypeError):
                    pass

        # Parse contest names from row5 raw text: "1 Name1     2 Name2"
        contest_name_map: dict[int, str] = {}
        # Split on pattern: "N ContestName" where N changes
        parts = re.split(r'\s{3,}', contest_raw)
        for part in parts:
            part = part.strip()
            m = re.match(r'^(\d+)\s+(.+)', part)
            if m:
                num = int(m.group(1))
                name = m.group(2).strip()
                contest_name_map[num] = name

        # If only one contest, it's simpler
        if not contest_nums:
            # All candidate columns belong to one contest
            contest_name = re.sub(r'^\d+\s+', '', contest_raw).strip()
            contest_nums = {1: list(range(7, len(row7)))}
            contest_name_map = {1: contest_name}

        # Extract registration/turnout from the Electionwide row
        reg_voters = _safe_int(row8[2]) if len(row8) > 2 else None
        voters_cast = _safe_int(row8[3]) if len(row8) > 3 else None

        # Process each contest group
        for num, col_indices in sorted(contest_nums.items()):
            contest_name = contest_name_map.get(num, f"Contest {num}")

            # Get or create race
            if contest_name in results_map:
                race = results_map[contest_name]
            else:
                race = ClarityRaceResult(
                    contest_name=contest_name,
                    is_measure=_is_measure(contest_name),
                    registered_voters=reg_voters,
                    ballots_cast=voters_cast,
                )
                if reg_voters and voters_cast:
                    race.turnout_pct = round(voters_cast / reg_voters * 100, 2)
                results_map[contest_name] = race

            # Extract candidates from row 7 and their votes from row 8
            for col_idx in col_indices:
                cand_name = str(row7[col_idx]).strip() if col_idx < len(row7) and row7[col_idx] else ""
                if not cand_name or cand_name in ('Registered Voters', 'Voters Cast', 'Turnout (%)'):
                    continue

                votes = _safe_int(row8[col_idx]) if col_idx < len(row8) else 0

                # Parse party from "PARTY - NAME" prefix format
                name, party = _split_party_prefix(cand_name)

                # Skip if this candidate already exists (from multi-sheet merge)
                existing = [c for c in race.candidates if c.name == _clean_name(name)]
                if existing:
                    continue

                race.candidates.append(ClarityCandidateResult(
                    name=_clean_name(name),
                    party=party,
                    votes=votes,
                    is_writein=_is_writein(cand_name),
                ))

    # Post-process: compute percentages, sort
    results: list[ClarityRaceResult] = []
    for race in results_map.values():
        total_votes = sum(c.votes for c in race.candidates)
        for c in race.candidates:
            if total_votes > 0:
                c.vote_pct = round(c.votes / total_votes * 100, 2)
        race.candidates.sort(key=lambda c: -c.votes)
        results.append(race)

    logger.info("District canvass parsed: %d contests", len(results))
    return results


def _split_party_prefix(text: str) -> tuple[str, str | None]:
    """Split 'DEM - JOE BIDEN/KAMALA HARRIS' into (name, party).

    Also handles 'P&F - GLORIA LA RIVA' and non-partisan names.
    """
    m = re.match(r'^([A-Z&]{2,5})\s+-\s+(.+)$', text)
    if m:
        party = m.group(1)
        name = m.group(2)
        # Map common abbreviations
        party_map = {'P&F': 'PF', 'P&amp;F': 'PF'}
        party = party_map.get(party, party)
        return name, party
    return text, None


# ─── Parser G/H: SOV by-District/Precinct (2020 Primary, 2018) ──

def _parse_sov(wb: openpyxl.Workbook) -> list[ClarityRaceResult]:
    """Parse Statement of Votes Cast format (2020 Primary, 2018 General/Primary).

    Key features:
      - Named contest sheets (Gov, DemPP, Prop13, etc.) or SheetN
      - Row 1: contest title with "(Vote for X)" suffix
      - Row 3: split-table header — left metadata, right candidates
      - Candidate names have embedded newlines: "NAME\\n(PARTY)"
      - Look for "- Total" rows for vote totals
    """
    results: list[ClarityRaceResult] = []
    skip_sheets = {'Election Summary Report', 'Precinct Summary', 'District Summary',
                   'Sheet1', 'Shasta Summary 6.2018'}

    def _cell(row, idx, default=""):
        """Safely get a cell value from a row tuple."""
        if not row or idx >= len(row) or row[idx] is None:
            return default
        return str(row[idx]).strip()

    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue

        try:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if len(rows) < 5:
                continue

            # Row 1 (idx 1): contest title
            contest_title = _cell(rows[1], 0)
            if not contest_title:
                continue

            # Strip "(Vote for X)" suffix
            contest_title = re.sub(r'\s*\(Vote\s+for\s+\d+\)\s*$', '', contest_title, flags=re.IGNORECASE).strip()

            # Row 3 (idx 3): header with candidate names
            if len(rows) <= 3:
                continue
            header = rows[3]
            if not header or len(header) < 4:
                continue

            # Find the RIGHT-side candidate columns
            right_start = None
            for ci in range(3, min(len(header), 10)):
                text = _cell(header, ci)
                if text.lower() in ('precinct', 'district'):
                    right_start = ci + 1
                    break

            if right_start is None:
                right_start = 4

            # Parse candidate columns from header
            candidate_cols: list[tuple[int, str, str | None]] = []
            ci = right_start
            while ci < len(header):
                text = _cell(header, ci)
                if not text:
                    ci += 1
                    continue
                if text.lower() in ('total votes', 'undervotes', 'overvotes', 'total'):
                    ci += 1
                    continue

                name, party = _parse_sov_candidate_name(text)
                if name and not _is_rejected_writein(name):
                    candidate_cols.append((ci, name, party))
                    ci += 2  # skip percentage column
                else:
                    ci += 1

            if not candidate_cols:
                continue

            # Find totals row: scan from bottom for "- Total" pattern
            totals_row = None
            for ri in range(len(rows) - 1, 3, -1):
                cell_a = _cell(rows[ri], 0)
                if cell_a.endswith('- Total'):
                    totals_row = rows[ri]
                    break

            if totals_row is None:
                # Try "Electionwide" row
                for ri in range(4, min(len(rows), 20)):
                    cell_a = _cell(rows[ri], 0)
                    if cell_a.lower() == 'electionwide':
                        totals_row = rows[ri]
                        break

            if totals_row is None:
                logger.debug("Skipping sheet '%s' — no totals row", sheet_name)
                continue

            reg_voters = _safe_int(_cell(totals_row, 2, "0"))
            times_cast = _safe_int(_cell(totals_row, 1, "0"))

            race = ClarityRaceResult(
                contest_name=contest_title,
                is_measure=_is_measure(contest_title),
                registered_voters=reg_voters if reg_voters else None,
                ballots_cast=times_cast if times_cast else None,
            )
            if reg_voters and times_cast:
                race.turnout_pct = round(times_cast / reg_voters * 100, 2)

            total_cast = 0
            for col_idx, name, party in candidate_cols:
                votes = _safe_int(_cell(totals_row, col_idx, "0"))
                total_cast += votes
                race.candidates.append(ClarityCandidateResult(
                    name=_clean_name(name),
                    party=party,
                    votes=votes,
                    is_writein=_is_writein(name),
                ))

            for c in race.candidates:
                if total_cast > 0:
                    c.vote_pct = round(c.votes / total_cast * 100, 2)

            race.candidates.sort(key=lambda c: -c.votes)
            results.append(race)

        except Exception as e:
            logger.warning("Error parsing sheet '%s': %s", sheet_name, e)
            continue

    logger.info("SOV parsed: %d contests from %d sheets", len(results), len(wb.sheetnames))
    return results


def _parse_sov_candidate_name(text: str) -> tuple[str, str | None]:
    """Parse a SOV header candidate name like 'JOHN H. COX\\n(REP) ' -> (name, party)."""
    if '\n' in text:
        parts = text.split('\n')
        name = parts[0].strip()
        rest = parts[1].strip() if len(parts) > 1 else ""
        # Extract (PARTY) from second line
        m = re.match(r'\(([A-Z]{2,5})\)', rest)
        if m:
            return name, m.group(1)
        # Handle "Qualified Write In" or similar
        if 'write' in rest.lower():
            return name + ' (W)', None
        return name, None
    return text.strip(), None


# ─── Parser C: CVR ballot-level CSV ──────────────────────────

def _parse_cvr_csv(path: Path) -> list[ClarityRaceResult]:
    """Parse Cast Vote Record (CVR) CSV — ballot-level data.

    3-row header:
      Row 0: contest names (e.g., "Governor (Vote For=1)") — repeated per candidate column
      Row 1: candidate names
      Row 2: party abbreviations (or metadata field names for cols 0-7)
    Data rows (3+): 0 or 1 per candidate per ballot.
    """
    logger.info("Parsing CVR: %s", path.name)

    # Read header rows — detect if there's an extra title row
    with open(path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        row0 = next(reader, [])
        row1 = next(reader, [])
        row2 = next(reader, [])
        row3 = next(reader, [])

    if len(row0) < 10:
        logger.warning("CVR file too few columns: %s", path.name)
        return []

    # Detect title row: if row0 has a non-contest string in first cell and
    # the contest names are actually in row1
    contest_row, candidate_row, party_row = row0, row1, row2
    header_rows_to_skip = 3

    # Check if row0 is a title (first cell non-empty, no "(Vote For=" patterns)
    if row0[0].strip() and '(Vote For=' not in row0[0] and '(Vote for=' not in row0[0]:
        # Check if row1 has contest names
        if any('(Vote For=' in c or '(Vote for=' in c for c in row1 if c.strip()):
            logger.info("CVR: detected title row, shifting header by 1")
            contest_row, candidate_row, party_row = row1, row2, row3
            header_rows_to_skip = 4

    # Determine metadata column count (cols 0-7 are typically metadata)
    meta_cols = 0
    for i, cell in enumerate(party_row):
        cell = cell.strip()
        if cell and cell in ('CvrNumber', 'TabulatorNum', 'BatchId', 'RecordId',
                              'ImprintedId', 'CountingGroup', 'PrecinctPortion', 'BallotType'):
            meta_cols = i + 1
        elif i > 10:
            break
    if meta_cols == 0:
        meta_cols = 8  # default

    # Build contest -> candidate mapping from header
    # contest_name -> [(col_idx, candidate_name, party)]
    contests: dict[str, list[tuple[int, str, str | None]]] = {}
    for col_idx in range(meta_cols, len(contest_row)):
        contest = contest_row[col_idx].strip() if col_idx < len(contest_row) else ""
        candidate = candidate_row[col_idx].strip() if col_idx < len(candidate_row) else ""
        party = party_row[col_idx].strip() if col_idx < len(party_row) else ""

        if not contest or not candidate:
            continue

        # Clean contest name — strip "(Vote For=N)" suffix
        clean_contest = re.sub(r'\s*\(Vote\s+For=\d+\)\s*$', '', contest, flags=re.IGNORECASE).strip()

        if not party or party in ('', 'None'):
            party = None

        contests.setdefault(clean_contest, []).append((col_idx, candidate, party))

    if not contests:
        logger.warning("CVR: no contests found in header: %s", path.name)
        return []

    # Initialize vote counters
    vote_counts: dict[str, dict[int, int]] = {}  # contest -> {col_idx: count}
    ballot_count = 0
    for contest_name, candidates in contests.items():
        vote_counts[contest_name] = {col_idx: 0 for col_idx, _, _ in candidates}

    # Stream through data rows to count votes
    with open(path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        # Skip header rows
        for _ in range(header_rows_to_skip):
            next(reader, None)

        for row in reader:
            ballot_count += 1
            for contest_name, candidates in contests.items():
                for col_idx, _, _ in candidates:
                    if col_idx < len(row):
                        try:
                            val = int(row[col_idx])
                            if val == 1:
                                vote_counts[contest_name][col_idx] += 1
                        except (ValueError, IndexError):
                            pass

    # Build results
    results: list[ClarityRaceResult] = []
    for contest_name, candidates in contests.items():
        race = ClarityRaceResult(
            contest_name=contest_name,
            is_measure=_is_measure(contest_name),
            ballots_cast=ballot_count,
        )

        total_votes = sum(vote_counts[contest_name].values())
        for col_idx, cand_name, party in candidates:
            votes = vote_counts[contest_name][col_idx]
            if _is_rejected_writein(cand_name):
                continue
            race.candidates.append(ClarityCandidateResult(
                name=_clean_name(cand_name),
                party=party,
                votes=votes,
                vote_pct=round(votes / total_votes * 100, 2) if total_votes > 0 else None,
                is_writein=_is_writein(cand_name),
            ))

        race.candidates.sort(key=lambda c: -c.votes)
        results.append(race)

    logger.info("CVR parsed: %d contests from %d ballots", len(results), ballot_count)
    return results


# ─── Generic Excel (older formats) ──────────────────────────

def _parse_generic_excel(wb: openpyxl.Workbook) -> list[ClarityRaceResult]:
    """Attempt to parse older Excel formats that don't match B1/B2 exactly.

    These include 2020 and 2021 "Statement of Votes Cast" files which may have
    different layouts. Returns empty list if format is unrecognized.
    """
    results: list[ClarityRaceResult] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 10:
            continue

        # Try to detect canvass-like format (similar to B2)
        # Look for "Totals" row and candidate headers
        header_row_idx = None
        totals_row_idx = None

        for ri, row in enumerate(rows):
            cell_a = str(row[0]).strip() if row[0] else ""
            if cell_a.lower() == 'precinct' and header_row_idx is None:
                header_row_idx = ri
            if cell_a.lower() == 'totals' and ri > 10:
                totals_row_idx = ri

        if header_row_idx is not None and totals_row_idx is not None:
            # Try canvass-like parsing for this sheet
            # Look for contest title above the header
            contest_name = ""
            for ri in range(max(0, header_row_idx - 10), header_row_idx):
                cell = rows[ri][0] if rows[ri][0] else ""
                text = str(cell).strip()
                if text and len(text) > 5 and not text.startswith(('Canvass', 'Statement')):
                    if not re.match(r'^\d{1,2}[/.-]\d{1,2}[/.-]\d{2,4}$', text):
                        contest_name = text

            if not contest_name:
                contest_name = sheet_name

            header_row = rows[header_row_idx]
            totals_row = rows[totals_row_idx]

            candidate_cols: list[tuple[int, str]] = []
            for col_idx, cell in enumerate(header_row):
                if cell is None:
                    continue
                text = str(cell).strip()
                if text and text.lower() not in ('precinct', 'precinct id', ''):
                    candidate_cols.append((col_idx, text))

            if candidate_cols:
                race = ClarityRaceResult(
                    contest_name=contest_name,
                    is_measure=_is_measure(contest_name),
                )

                total_cast = 0
                for col_idx, cand_name in candidate_cols:
                    votes = totals_row[col_idx] if col_idx < len(totals_row) else 0
                    votes = int(votes) if votes else 0
                    total_cast += votes

                    name, party = _split_party(cand_name)
                    race.candidates.append(ClarityCandidateResult(
                        name=_clean_name(name),
                        party=party,
                        votes=votes,
                        is_writein=_is_writein(cand_name),
                    ))

                for c in race.candidates:
                    if total_cast > 0:
                        c.vote_pct = round(c.votes / total_cast * 100, 2)

                race.candidates.sort(key=lambda c: -c.votes)
                results.append(race)

    logger.info("Generic Excel parsed: %d contests from %d sheets",
                len(results), len(wb.sheetnames))
    return results


# ─── Helpers ─────────────────────────────────────────────────

def _safe_int(val: str | int | float | None) -> int:
    """Safely convert to int, returning 0 on failure."""
    if val is None:
        return 0
    try:
        return int(float(str(val).replace(',', '').strip()))
    except (ValueError, TypeError):
        return 0


def _is_measure(contest_name: str) -> bool:
    """Check if a contest is a ballot measure."""
    name = contest_name.lower().strip()
    return bool(re.match(
        r'^(measure|prop|proposition|recall)\b', name,
    ))


def _is_writein(name: str) -> bool:
    """Check if a candidate is a write-in."""
    lower = name.lower()
    return '(w)' in lower or 'write-in' in lower or 'write in' in lower


def _is_rejected_writein(name: str) -> bool:
    """Check if this is a rejected/unassigned write-in noise entry."""
    lower = name.lower()
    return any(kw in lower for kw in [
        'rejected write', 'unassigned write', 'unresolved write',
    ])


def _clean_name(name: str) -> str:
    """Clean candidate name — strip (W) suffix, normalize whitespace."""
    name = re.sub(r'\s*\(W\)\s*$', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s+', ' ', name).strip()
    return name


def _split_party(text: str) -> tuple[str, str | None]:
    """Split 'CANDIDATE NAME, REP' into (name, party).

    Common party abbreviations: REP, DEM, AIP, LIB, GRN, PF, NPP
    """
    party_abbrevs = {'REP', 'DEM', 'AIP', 'LIB', 'GRN', 'PF', 'NPP', 'AI'}

    # Check for trailing comma + party abbreviation
    match = re.match(r'^(.+?),\s*(\w{2,3})\s*$', text)
    if match and match.group(2).upper() in party_abbrevs:
        return match.group(1).strip(), match.group(2).upper()

    return text, None
